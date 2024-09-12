import pandas as pd

import tkinter as tk

from tkinter import filedialog, messagebox

import os

from openpyxl import Workbook

from openpyxl.utils.dataframe import dataframe_to_rows

from openpyxl.styles import Alignment, Font




def process_master_csv(master_file_path, low_stock_threshold):

    # Read the master CSV file

    master_df = pd.read_csv(master_file_path)




    # Filter for Front of House and Back of House

    front_of_house_df = master_df[master_df['Room'] == 'Sales Floor']

    back_of_house_df = master_df[master_df['Room'] == 'Vault']




    # Process the data as previously done

    return process_csv_files(front_of_house_df, back_of_house_df, low_stock_threshold)




def process_csv_files(front_of_house_df, back_of_house_df, low_stock_threshold):

    # Initialize an empty DataFrame to store low stock items

    low_stock_items_list = []




    # Convert 'Available' column to float

    front_of_house_df['Available'] = front_of_house_df['Available'].astype(float)

    

    # Group by Brand, Online title and sum the quantities

    grouped_chunk = front_of_house_df.groupby(['Brand', 'Online title'])['Available'].sum().reset_index()

    

    # Identify low stock items

    low_stock_chunk = grouped_chunk[grouped_chunk['Available'] < low_stock_threshold]

        

    # Append to the list

    low_stock_items_list.append(low_stock_chunk)




    # Concatenate all low stock items

    low_stock_items = pd.concat(low_stock_items_list).groupby(['Brand', 'Online title'])['Available'].sum().reset_index()




    # Convert 'Available' column to float

    back_of_house_df['Available'] = back_of_house_df['Available'].astype(float)




    # Check back of house availability

    low_stock_items = low_stock_items.merge(back_of_house_df[['Brand', 'Online title', 'Available']],

                                            on=['Brand', 'Online title'], how='left', suffixes=('_front', '_back'))




    # Rename columns for clarity

    low_stock_items = low_stock_items.rename(columns={'Available_front': 'Available Front', 'Available_back': 'Available Back'})




    # Filter items available in back of house

    available_in_back = low_stock_items[low_stock_items['Available Back'] > 0]




    return available_in_back




   




def select_file(entry):

    file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])

    entry.delete(0, tk.END)

    entry.insert(0, file_path)




def run_processing():

    master_file_path = master_entry.get()

    low_stock_threshold = threshold_entry.get()

    

    if not master_file_path or not low_stock_threshold:

        messagebox.showerror("Error", "Please select the master CSV file and set the low stock threshold.")

        return




    try:

        low_stock_threshold = float(low_stock_threshold)

    except ValueError:

        messagebox.showerror("Error", "Low stock threshold must be a number.")

        return

    

    try:

        # Process CSV files

        low_stock_items_df = process_master_csv(master_file_path, low_stock_threshold)

        

        # Clear any existing content in the text widget

        report_text.delete(1.0, tk.END)

        

        # Convert DataFrame to a formatted string and display in the text widget

        report_text.insert(tk.END, format_dataframe(low_stock_items_df))

        

        # Save the report to a new Excel file with gridlines and autofit column widths

        excel_file = 'low_stock_report.xlsx'

        save_to_excel_with_autofit_and_gridlines(low_stock_items_df, excel_file)

        

        print(f"Report generated: {excel_file}")

        

        messagebox.showinfo("Success", f"Report generated successfully: {excel_file}")

    except Exception as e:

        messagebox.showerror("Error", f"An error occurred: {e}")




def save_to_excel_with_autofit_and_gridlines(df, filename):

    from openpyxl import Workbook

    from openpyxl.worksheet.page import PageMargins, PrintOptions




    wb = Workbook()

    ws = wb.active

    

    # Add a heading

    heading = ["What's Missing"]

    ws.append(heading)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=df.shape[1])

    ws['A1'].font = Font(size=14, bold=True)

    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    

    # Write the DataFrame to the worksheet

    for r in dataframe_to_rows(df, index=False, header=True):

        ws.append(r)

    

    # Set alignment to center for all cells

    alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=2):

        for cell in row:

            cell.alignment = alignment

    

    # Auto-fit column widths, skipping the merged header cell

    for col in ws.iter_cols(min_row=2, max_col=ws.max_column):

        max_length = 0

        column = col[0].column_letter

        for cell in col:

            try:

                if len(str(cell.value)) > max_length:

                    max_length = len(cell.value)

            except:

                pass

        adjusted_width = (max_length + 2)

        ws.column_dimensions[column].width = adjusted_width

    

    # Enable gridlines and set print orientation to landscape

    ws.sheet_view.showGridLines = True

    ws.print_options = PrintOptions(horizontalCentered=True, verticalCentered=False, gridLines=True)

    ws.page_setup.orientation = 'landscape'

    

    # Set margins for better printing layout

    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75, header=0.3, footer=0.3)

    

    # Save the Excel file

    wb.save(filename)










    




def print_excel():

    try:

        os.startfile('low_stock_report.xlsx', 'print')

        messagebox.showinfo("Success", "Excel file sent to printer successfully.")

    except Exception as e:

        messagebox.showerror("Error", f"An error occurred while printing: {e}")




def format_dataframe(df):

    # Convert DataFrame to a formatted string for display

    formatted_string = df.to_string(index=False)

    return formatted_string




# Setup GUI

root = tk.Tk()

root.title("What's Missing")




# Font configuration for the report text

text_font = ("Courier", 10)  # Fixed-width font




# Master CSV section

master_label = tk.Label(root, text="Select Valuation Report")

master_label.pack()

master_entry = tk.Entry(root, width=50)

master_entry.pack()

master_button = tk.Button(root, text="Browse", command=lambda: select_file(master_entry))

master_button.pack()




# Low Stock Threshold section

threshold_label = tk.Label(root, text="Set Low Stock Threshold")

threshold_label.pack()

threshold_entry = tk.Entry(root, width=50)

threshold_entry.pack()




# Generate Report button

generate_button = tk.Button(root, text="Generate Report", command=run_processing)

generate_button.pack()




# Print Excel button

print_button = tk.Button(root, text="Print", command=print_excel)

print_button.pack()




# Scrollbar for the report text widget

scrollbar = tk.Scrollbar(root)

scrollbar.pack(side=tk.RIGHT, fill=tk.Y)




# Text widget to display the report

report_text = tk.Text(root, height=20, width=120, wrap=tk.NONE, font=text_font, yscrollcommand=scrollbar.set, xscrollcommand=scrollbar.set)

report_text.pack()




# Configure the scrollbar

scrollbar.config(command=report_text.yview)




root.mainloop()
